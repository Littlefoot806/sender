from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC 
from time import sleep
from openpyxl import workbook, load_workbook
import json

def start_webdriver():
    driver = webdriver.Chrome()
    driver.get("https://accounts.google.com/ServiceLogin/signinchooser?elo=1&flowName=GlifWebSignIn&flowEntry=ServiceLogin")
    sleep(2)
    return driver

def loggining(driver, login, passwd):
    try:
        login_elem = driver.find_element_by_xpath('//input[@type="email"]')
        login_elem.clear()
        login_elem.send_keys(login)
        login_elem.send_keys(Keys.RETURN)

        sleep(3)

        pass_elem = driver.find_element_by_xpath('//input[@type="password"]')
        pass_elem.clear()
        pass_elem.send_keys(passwd)
        pass_elem.send_keys(Keys.RETURN)
        sleep(2)
        return True
    except Exception as e:
        print(e)
        return False

def get_emails():
	emails = []
	filename = 'list_of_email.xlsx'
	wb = load_workbook(filename)
	ws = wb.active
	for i in range(1,50):
		if ws['B{}'.format(i)].value == '+':
			continue
		emails.append(ws['A{}'.format(i)].value)
		ws['B{}'.format(i)] = '+'	
		if len(emails) >= 15:
			break
		
	print(emails)	
	wb.save(filename=filename)
	return emails
	
def send_mail(driver, emails, letter):
	
	try:
		click_write = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//div[@role="button" and contains(text(), "Написать")]')))
		click_write.click()
	except Exception as e:
		print(e)
		
	theme = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//input[@name="subjectbox"]')))
	body = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//div[@aria-label="Тело письма"]')))
	
	for email in emails:
		to = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//textarea[@name="to"]')))
		to.send_keys(email+' ')
	theme.send_keys(letter['theme'])
	body.send_keys(letter['body'])
	
	button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//div[contains(@aria-label,"Отправить") and @role="button"]')))
	button.click()
	sleep(10)
	
def main():
	login = "example@gmail.com" # google email
	passwd = "passExample"      # google password
	print('Starting...')


	driver = start_webdriver()
	if loggining(driver, login, passwd):
		driver.get('https://mail.google.com/mail/ca/u/0/')
		sleep(10)
		
		emails = get_emails()
		
		with open('letter.json', 'r') as f:
			letter = json.load(f)
			print(letter['theme'])

		send_mail(driver, emails, letter)
	else:
		print('Failed Authorization')
		driver.close()


if __name__ == '__main__':
    main()
	
	
	
	
	
	
	
	
	
	
	
	
	
